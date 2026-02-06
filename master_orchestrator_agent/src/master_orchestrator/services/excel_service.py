"""Excel Service for generating Excel output files."""

from pathlib import Path

import structlog
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

from master_orchestrator.models.unified_result import MasterAnalysisResult

logger = structlog.get_logger(__name__)


class ExcelService:
    """Service for generating Excel output files."""

    def __init__(self):
        """Initialize the Excel service."""
        # Define styles
        self._header_font = Font(bold=True, color="FFFFFF")
        self._header_fill = PatternFill(
            start_color="4472C4", end_color="4472C4", fill_type="solid"
        )
        self._border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self._center_alignment = Alignment(horizontal="center", vertical="center")
        self._wrap_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)

    def generate(self, result: MasterAnalysisResult, output_path: Path) -> None:
        """Generate Excel workbook from analysis result.

        Args:
            result: MasterAnalysisResult to convert
            output_path: Path to save the Excel file
        """
        logger.info("generating_excel", path=str(output_path))

        wb = Workbook()

        # Remove default sheet
        default_sheet = wb.active
        if default_sheet:
            wb.remove(default_sheet)

        # Create sheets for each section
        self._create_summary_sheet(wb, result)
        self._create_passport_sheet(wb, result)
        self._create_education_sheet(wb, result)
        self._create_financial_sheet(wb, result)
        self._create_cross_validation_sheet(wb, result)

        # Save workbook
        wb.save(output_path)
        logger.info("excel_generated", path=str(output_path))

    def _create_summary_sheet(self, wb: Workbook, result: MasterAnalysisResult) -> None:
        """Create summary overview sheet."""
        ws = wb.create_sheet("Summary")

        # Title
        ws["A1"] = "Document Analysis Summary"
        ws["A1"].font = Font(bold=True, size=14)
        ws.merge_cells("A1:D1")

        # Summary data
        data = [
            ["Section", "Status", "Key Finding"],
            [
                "Passport",
                result.passport_details.extraction_status if result.passport_details else "N/A",
                (
                    f"{result.passport_details.full_name or 'N/A'}"
                    if result.passport_details
                    else "Not processed"
                ),
            ],
            [
                "Education",
                result.education_summary.extraction_status if result.education_summary else "N/A",
                (
                    result.education_summary.validation_status.value
                    if result.education_summary
                    else "Not processed"
                ),
            ],
            [
                "Financial",
                result.financial_summary.extraction_status if result.financial_summary else "N/A",
                (
                    result.financial_summary.worthiness_status.value
                    if result.financial_summary
                    else "Not processed"
                ),
            ],
            [
                "Cross-Validation",
                "Completed" if result.cross_validation else "N/A",
                (
                    f"Name: {result.cross_validation.name_match}, DOB: {result.cross_validation.dob_match}"
                    if result.cross_validation
                    else "Not performed"
                ),
            ],
        ]

        self._write_table(ws, data, start_row=3)

        # Processing metadata
        if result.metadata:
            ws["A10"] = "Processing Metadata"
            ws["A10"].font = Font(bold=True)

            meta_data = [
                ["Metric", "Value"],
                ["Documents Scanned", str(result.metadata.total_documents_scanned)],
                [
                    "Processing Time",
                    f"{result.metadata.processing_time_seconds:.2f}s"
                    if result.metadata.processing_time_seconds
                    else "N/A",
                ],
                ["Errors", str(len(result.metadata.processing_errors))],
                ["Warnings", str(len(result.metadata.processing_warnings))],
            ]

            self._write_table(ws, meta_data, start_row=11)

        self._auto_width(ws)

    def _create_passport_sheet(self, wb: Workbook, result: MasterAnalysisResult) -> None:
        """Create passport details sheet."""
        ws = wb.create_sheet("Passport")

        ws["A1"] = "Passport Details"
        ws["A1"].font = Font(bold=True, size=14)

        if not result.passport_details:
            ws["A3"] = "No passport data available"
            return

        pd = result.passport_details
        data = [
            ["Field", "Value"],
            ["First Name", pd.first_name or ""],
            ["Last Name", pd.last_name or ""],
            ["Date of Birth", pd.date_of_birth or ""],
            ["Sex", pd.sex or ""],
            ["Passport Number", pd.passport_number or ""],
            ["Issuing Country", pd.issuing_country or ""],
            ["Issue Date", pd.issue_date or ""],
            ["Expiry Date", pd.expiry_date or ""],
            ["Accuracy Score", str(pd.accuracy_score)],
            ["Extraction Status", pd.extraction_status],
        ]

        if pd.mrz_data:
            data.extend(
                [
                    ["", ""],  # Empty row
                    ["MRZ Data", ""],
                    ["Document Type", pd.mrz_data.document_type or ""],
                    ["Raw Line 1", pd.mrz_data.raw_line1 or ""],
                    ["Raw Line 2", pd.mrz_data.raw_line2 or ""],
                    [
                        "Checksum Valid",
                        str(pd.mrz_data.checksum_valid) if pd.mrz_data.checksum_valid is not None else "N/A",
                    ],
                ]
            )

        self._write_table(ws, data, start_row=3)
        self._auto_width(ws)

    def _create_education_sheet(self, wb: Workbook, result: MasterAnalysisResult) -> None:
        """Create education summary sheet."""
        ws = wb.create_sheet("Education")

        ws["A1"] = "Education Summary"
        ws["A1"].font = Font(bold=True, size=14)

        if not result.education_summary:
            ws["A3"] = "No education data available"
            return

        ed = result.education_summary
        data = [
            ["Field", "Value"],
            ["Highest Qualification", ed.highest_qualification or ""],
            ["Institution", ed.institution or ""],
            ["Country", ed.country or ""],
            ["Student Name", ed.student_name or ""],
            ["Final Grade (Original)", ed.final_grade_original or ""],
            [
                "French Equivalent (0-20)",
                str(ed.french_equivalent_grade_0_20) if ed.french_equivalent_grade_0_20 else "N/A",
            ],
            ["Validation Status", ed.validation_status.value],
            ["Remarks", ed.remarks or ""],
            ["Extraction Status", ed.extraction_status],
        ]

        self._write_table(ws, data, start_row=3)
        self._auto_width(ws)

    def _create_financial_sheet(self, wb: Workbook, result: MasterAnalysisResult) -> None:
        """Create financial summary sheet."""
        ws = wb.create_sheet("Financial")

        ws["A1"] = "Financial Summary"
        ws["A1"].font = Font(bold=True, size=14)

        if not result.financial_summary:
            ws["A3"] = "No financial data available"
            return

        fs = result.financial_summary
        data = [
            ["Field", "Value"],
            ["Document Type", fs.document_type or ""],
            ["Account Holder", fs.account_holder_name or ""],
            ["Bank Name", fs.bank_name or ""],
            ["Base Currency", fs.base_currency or ""],
            ["Amount (Original)", str(fs.amount_original) if fs.amount_original else "N/A"],
            ["Amount (EUR)", str(fs.amount_eur) if fs.amount_eur else "N/A"],
            ["Threshold (EUR)", str(fs.financial_threshold_eur)],
            ["Worthiness Status", fs.worthiness_status.value],
            ["Remarks", fs.remarks or ""],
            ["Extraction Status", fs.extraction_status],
        ]

        self._write_table(ws, data, start_row=3)
        self._auto_width(ws)

    def _create_cross_validation_sheet(self, wb: Workbook, result: MasterAnalysisResult) -> None:
        """Create cross-validation sheet."""
        ws = wb.create_sheet("Cross-Validation")

        ws["A1"] = "Cross-Validation Results"
        ws["A1"].font = Font(bold=True, size=14)

        if not result.cross_validation:
            ws["A3"] = "No cross-validation data available"
            return

        cv = result.cross_validation
        data = [
            ["Check", "Result", "Details"],
            [
                "Name Match",
                str(cv.name_match) if cv.name_match is not None else "N/A",
                f"Score: {cv.name_match_score:.2f}" if cv.name_match_score else "N/A",
            ],
            [
                "DOB Match",
                str(cv.dob_match) if cv.dob_match is not None else "N/A",
                "",
            ],
            ["", "", ""],
            ["Remarks", cv.remarks or "", ""],
        ]

        self._write_table(ws, data, start_row=3)

        # Add comparison details
        ws["A10"] = "Comparison Details"
        ws["A10"].font = Font(bold=True)

        detail_data = [
            ["Source", "Name", "DOB"],
            ["Passport", cv.passport_name or "N/A", cv.passport_dob or "N/A"],
            ["Education", cv.education_name or "N/A", cv.education_dob or "N/A"],
            ["Financial", cv.financial_name or "N/A", "N/A"],
        ]

        self._write_table(ws, detail_data, start_row=11)
        self._auto_width(ws)

    def _write_table(
        self,
        ws,
        data: list[list[str]],
        start_row: int = 1,
        start_col: int = 1,
    ) -> None:
        """Write a table with headers to worksheet.

        Args:
            ws: Worksheet to write to
            data: List of rows, first row is header
            start_row: Starting row number
            start_col: Starting column number
        """
        for row_idx, row in enumerate(data):
            for col_idx, value in enumerate(row):
                cell = ws.cell(
                    row=start_row + row_idx,
                    column=start_col + col_idx,
                    value=value,
                )

                # Apply border
                cell.border = self._border

                # Apply header styling to first row
                if row_idx == 0:
                    cell.font = self._header_font
                    cell.fill = self._header_fill
                    cell.alignment = self._center_alignment
                else:
                    cell.alignment = self._wrap_alignment

    def _auto_width(self, ws, min_width: int = 10, max_width: int = 50) -> None:
        """Auto-adjust column widths based on content.

        Args:
            ws: Worksheet to adjust
            min_width: Minimum column width
            max_width: Maximum column width
        """
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                try:
                    cell_length = len(str(cell.value or ""))
                    max_length = max(max_length, cell_length)
                except Exception:
                    pass

            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width
